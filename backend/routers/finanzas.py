import unicodedata
from datetime import date, datetime, timedelta
from typing import List, Optional
from collections import defaultdict
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from database import get_db
from fechas import TZ_BOGOTA, fin_dia_utc, hoy_bogota, inicio_dia_utc
from models import Producto, MovimientoFinanciero, Pago, Plan, RolUsuario, TipoMovimiento, Usuario, Venta
from schemas.finanza import BalanceResponse, MovimientoCreate
from security import get_current_user

router = APIRouter(prefix="/finanzas", tags=["Finanzas"])


def _require_admin(current_user: Usuario = Depends(get_current_user)):
    if current_user.rol != RolUsuario.ADMIN:
        raise HTTPException(status_code=403, detail="Solo el administrador puede acceder al módulo financiero.")
    return current_user


def _apply_date_filter(query, model_fecha_col, desde: Optional[date], hasta: Optional[date]):
    """Acota a los días del NEGOCIO (Bogotá), no a los días UTC.

    `desde`/`hasta` llegan como fechas de Bogotá, pero las columnas se guardan naive en
    UTC. Acá se armaba la ventana con `datetime.combine(desde, min.time())`, que es
    medianoche **UTC** — las 19:00 de Bogotá del día anterior. Con el box entrenando de
    19:00 a 21:00, "Hoy" arrastraba la noche de ayer y perdía la de hoy.

    Los helpers viven en `fechas.py` y los comparten `dashboard` y `asistencia`; tenerlos
    duplicados fue exactamente la causa de este mismo bug allá. No los redefinas.
    """
    if desde:
        query = query.filter(model_fecha_col >= inicio_dia_utc(desde))
    if hasta:
        query = query.filter(model_fecha_col <= fin_dia_utc(hasta))
    return query


@router.get("/balance", response_model=BalanceResponse)
def balance(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin),
):
    # ── Ingresos de membresías (tabla pagos) — sumado en SQL ──
    q_pagos = db.query(func.coalesce(func.sum(Pago.monto), 0.0))
    q_pagos = _apply_date_filter(q_pagos, Pago.fecha_pago, fecha_desde, fecha_hasta)
    total_membresias = q_pagos.scalar() or 0.0

    # ── Ingresos de tienda (tabla ventas) — sumado en SQL ──
    q_ventas = db.query(func.coalesce(func.sum(Venta.total), 0.0))
    q_ventas = _apply_date_filter(q_ventas, Venta.fecha_venta, fecha_desde, fecha_hasta)
    total_ventas = q_ventas.scalar() or 0.0

    # ── Movimientos manuales — agrupados por (tipo, categoría) en SQL ──
    q_mov = db.query(
        MovimientoFinanciero.tipo,
        MovimientoFinanciero.categoria,
        func.coalesce(func.sum(MovimientoFinanciero.monto), 0.0),
    )
    q_mov = _apply_date_filter(q_mov, MovimientoFinanciero.fecha, fecha_desde, fecha_hasta)
    q_mov = q_mov.group_by(MovimientoFinanciero.tipo, MovimientoFinanciero.categoria)

    ingresos_manuales = defaultdict(float)
    egresos_por_categoria = defaultdict(float)
    for _tipo, _categoria, _suma in q_mov.all():
        if _tipo == TipoMovimiento.INGRESO:
            ingresos_manuales[_categoria] += _suma
        elif _tipo == TipoMovimiento.EGRESO:
            egresos_por_categoria[_categoria] += _suma

    total_ingresos_manuales = sum(ingresos_manuales.values())
    total_egresos = sum(egresos_por_categoria.values())

    ingresos_por_categoria = {
        "mensualidad": total_membresias + ingresos_manuales.get("mensualidad", 0),
        "venta_tienda": total_ventas,
        "ingreso_varios": ingresos_manuales.get("ingreso_varios", 0),
    }
    # Add any other manual ingreso categories
    for cat, val in ingresos_manuales.items():
        if cat not in ingresos_por_categoria:
            ingresos_por_categoria[cat] = val

    total_ingresos = total_membresias + total_ventas + total_ingresos_manuales

    return BalanceResponse(
        ingresos_total=round(total_ingresos, 2),
        total_membresias=round(total_membresias + ingresos_manuales.get("mensualidad", 0), 2),
        total_tienda=round(total_ventas, 2),
        egresos_total=round(total_egresos, 2),
        balance_neto=round(total_ingresos - total_egresos, 2),
        ingresos_por_categoria={k: round(v, 2) for k, v in ingresos_por_categoria.items() if v > 0},
        egresos_por_categoria={k: round(v, 2) for k, v in egresos_por_categoria.items()},
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )


def _sin_acentos(s: Optional[str]) -> str:
    """Minúsculas y sin tildes, para que el buscador no dependa de cómo se tipeó.

    Sin esto, buscar "nomina" no encuentra "Nómina" y el buscador parece roto: quien
    busca no tiene por qué saber con qué acento se cargó el concepto.
    """
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )


def _recolectar_movimientos(
    db: Session,
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    tipo: Optional[str] = None,
    q: Optional[str] = None,
    categoria: Optional[str] = None,
    plan_id: Optional[int] = None,
    metodo_pago: Optional[str] = None,
) -> List[dict]:
    """Unifica pagos + ventas + movimientos manuales del período, ordenados por fecha desc.

    Lo consumen el listado y la exportación a Excel: si la fusión de las tres
    fuentes se duplicara, el Excel y la pantalla podrían mostrar cosas distintas. Por
    eso los filtros viven **acá adentro** y no en el llamador — así el archivo exportado
    dice exactamente lo que el admin tiene en pantalla.

    `plan_id` solo tiene sentido sobre pagos de membresía: al usarlo se excluyen las
    ventas de tienda y los movimientos manuales, que no pertenecen a ningún plan.
    """
    items = []

    # Un filtro por plan o por una categoría que no sea de pagos deja fuera todo lo demás:
    # se evalúa antes de consultar para no traer filas que se van a descartar igual.
    quiere_pagos  = tipo in (None, "ingreso") and categoria in (None, "mensualidad")
    quiere_ventas = tipo in (None, "ingreso") and categoria in (None, "venta_tienda") and plan_id is None
    quiere_manual = plan_id is None

    # ── Pagos de membresías ──
    if quiere_pagos:
        # joinedload puebla plan y usuario en la misma query (evita lazy load por fila).
        q_pagos = db.query(Pago).options(joinedload(Pago.plan), joinedload(Pago.usuario))
        q_pagos = _apply_date_filter(q_pagos, Pago.fecha_pago, fecha_desde, fecha_hasta)
        if plan_id is not None:
            q_pagos = q_pagos.filter(Pago.plan_id == plan_id)
        if metodo_pago:
            q_pagos = q_pagos.filter(Pago.metodo_pago == metodo_pago)
        for p in q_pagos.order_by(Pago.fecha_pago.desc()).all():
            plan_nombre = p.plan.nombre if p.plan else "Personalizado"
            usuario_nombre = p.usuario.nombre if p.usuario else None
            items.append({
                "id": f"pago_{p.id}",
                "tipo": "ingreso",
                "concepto": f"Membresía – {usuario_nombre or 'Anónimo'} ({plan_nombre})",
                "categoria": "mensualidad",
                "monto": p.monto,
                "fecha": p.fecha_pago,
                "metodo_pago": p.metodo_pago,
                "usuario_nombre": usuario_nombre,
                "fuente": "pago_membresia",
                "es_eliminable": False,
            })

    # ── Ventas de tienda ──
    if quiere_ventas:
        q_ventas = db.query(Venta).options(joinedload(Venta.producto), joinedload(Venta.usuario))
        q_ventas = _apply_date_filter(q_ventas, Venta.fecha_venta, fecha_desde, fecha_hasta)
        if metodo_pago:
            q_ventas = q_ventas.filter(Venta.metodo_pago == metodo_pago)
        for v in q_ventas.order_by(Venta.fecha_venta.desc()).all():
            items.append({
                "id": f"venta_{v.id}",
                "tipo": "ingreso",
                "concepto": f"Venta tienda – {v.producto.nombre if v.producto else 'Producto'} ×{v.cantidad}",
                "categoria": "venta_tienda",
                "monto": v.total,
                "fecha": v.fecha_venta,
                "metodo_pago": v.metodo_pago,
                "usuario_nombre": v.usuario.nombre if v.usuario else None,
                "fuente": "venta_tienda",
                "es_eliminable": False,
            })

    # ── Movimientos manuales ──
    if not quiere_manual:
        q_mov_rows = []
    else:
        q_mov = db.query(MovimientoFinanciero).options(joinedload(MovimientoFinanciero.usuario))
        q_mov = _apply_date_filter(q_mov, MovimientoFinanciero.fecha, fecha_desde, fecha_hasta)
        if tipo in ("ingreso", "egreso"):
            q_mov = q_mov.filter(MovimientoFinanciero.tipo == TipoMovimiento(tipo))
        if categoria:
            q_mov = q_mov.filter(MovimientoFinanciero.categoria == categoria)
        if metodo_pago:
            q_mov = q_mov.filter(MovimientoFinanciero.metodo_pago == metodo_pago)
        q_mov_rows = q_mov.order_by(MovimientoFinanciero.fecha.desc()).all()
    for m in q_mov_rows:
        items.append({
            "id": f"mov_{m.id}",
            "tipo": m.tipo.value,
            "concepto": m.concepto,
            "categoria": m.categoria,
            "monto": m.monto,
            "fecha": m.fecha,
            "metodo_pago": m.metodo_pago,
            "usuario_nombre": m.usuario.nombre if m.usuario else None,
            "fuente": m.fuente,
            "es_eliminable": True,
        })

    # El buscador va al final y en memoria: pega contra el concepto y el nombre del
    # cliente, que en pagos y ventas son texto armado acá (no columnas), así que en SQL
    # habría que reconstruirlo por fuente y las tres consultas dejarían de coincidir.
    if q and q.strip():
        aguja = _sin_acentos(q)
        items = [
            it for it in items
            if aguja in _sin_acentos(it["concepto"]) or aguja in _sin_acentos(it["usuario_nombre"])
        ]

    items.sort(key=lambda x: x["fecha"], reverse=True)
    return items


@router.get("/movimientos")
def listar_movimientos(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    tipo: Optional[str] = Query(None),
    q: Optional[str] = Query(None, max_length=100, description="Busca en concepto y cliente"),
    categoria: Optional[str] = Query(None),
    plan_id: Optional[int] = Query(None),
    metodo_pago: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(15, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin),
):
    """Devuelve `{items, total}` — el total es del período filtrado, no de la página.

    La paginación es de servidor porque el buscador tiene que encontrar cualquier
    movimiento del período: recortando a 200 en el cliente, buscar algo que SÍ está y no
    aparezca es peor que no tener buscador.
    """
    todos = _recolectar_movimientos(
        db, fecha_desde, fecha_hasta, tipo,
        q=q, categoria=categoria, plan_id=plan_id, metodo_pago=metodo_pago,
    )
    return {
        "items": todos[skip:skip + limit],
        "total": len(todos),
        "total_monto": round(sum(m["monto"] for m in todos), 2),
    }


@router.get("/estadisticas")
def estadisticas(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin),
):
    """Qué se vendió en el período: desglose por plan y por producto.

    El balance dice cuánta plata entró; esto dice de dónde. Las dos agregaciones van en
    SQL con `group_by` + `outerjoin` para traer los nombres sin una consulta por fila.
    """
    # ── Planes ──
    # outerjoin porque un pago personalizado tiene `plan_id = NULL` y no matchea ninguna
    # fila de `planes`: con un join normal desaparecería del desglose, y es ingreso real.
    q_planes = (
        db.query(
            Pago.plan_id,
            Plan.nombre,
            func.count(Pago.id),
            func.coalesce(func.sum(Pago.monto), 0.0),
        )
        .outerjoin(Plan, Plan.id == Pago.plan_id)
    )
    q_planes = _apply_date_filter(q_planes, Pago.fecha_pago, fecha_desde, fecha_hasta)
    filas_planes = q_planes.group_by(Pago.plan_id, Plan.nombre).all()

    filas = [
        {"plan_id": plan_id, "nombre": nombre or "Personalizado", "cantidad": cantidad, "total": total}
        for plan_id, nombre, cantidad, total in filas_planes
    ]

    # Los ingresos manuales categorizados como "mensualidad" no tienen plan detrás, pero
    # la tarjeta "Membresías" del balance SÍ los suma. Sin esta fila, la tabla y la
    # tarjeta muestran números distintos uno al lado del otro y el admin no tiene cómo
    # saber cuál creer — que en una pantalla de plata es lo peor que puede pasar.
    q_manual = db.query(
        func.count(MovimientoFinanciero.id),
        func.coalesce(func.sum(MovimientoFinanciero.monto), 0.0),
    ).filter(
        MovimientoFinanciero.tipo == TipoMovimiento.INGRESO,
        MovimientoFinanciero.categoria == "mensualidad",
    )
    q_manual = _apply_date_filter(q_manual, MovimientoFinanciero.fecha, fecha_desde, fecha_hasta)
    cant_manual, total_manual = q_manual.one()
    if cant_manual:
        filas.append({
            "plan_id": None, "nombre": "Ingreso manual (sin plan)",
            "cantidad": cant_manual, "total": total_manual,
        })

    total_planes = sum(f["total"] for f in filas) or 0.0
    planes = [
        {
            **f,
            "total": round(f["total"], 2),
            "porcentaje": round(f["total"] / total_planes * 100, 1) if total_planes else 0.0,
        }
        for f in filas
    ]
    planes.sort(key=lambda p: p["total"], reverse=True)

    # ── Productos ──
    q_prod = (
        db.query(
            Venta.producto_id,
            Producto.nombre,
            func.coalesce(func.sum(Venta.cantidad), 0),
            func.coalesce(func.sum(Venta.total), 0.0),
        )
        .outerjoin(Producto, Producto.id == Venta.producto_id)
    )
    q_prod = _apply_date_filter(q_prod, Venta.fecha_venta, fecha_desde, fecha_hasta)
    filas_prod = q_prod.group_by(Venta.producto_id, Producto.nombre).all()

    total_prod = sum(f[3] for f in filas_prod) or 0.0
    productos = [
        {
            "producto_id": producto_id,
            "nombre": nombre or "Producto eliminado",
            "unidades": unidades,
            "total": round(total, 2),
            "porcentaje": round(total / total_prod * 100, 1) if total_prod else 0.0,
        }
        for producto_id, nombre, unidades, total in filas_prod
    ]
    productos.sort(key=lambda p: p["total"], reverse=True)

    return {
        "planes": planes,
        "productos": productos,
        "total_planes": round(total_planes, 2),
        "total_productos": round(total_prod, 2),
    }


_LABELS_CATEGORIA = {
    "mensualidad": "Membresía", "venta_tienda": "Tienda", "ingreso_varios": "Varios",
    "renta": "Renta", "servicios": "Servicios", "equipamiento": "Equipamiento",
    "nomina": "Nómina", "marketing": "Marketing", "mantenimiento": "Mantenimiento",
    "otros": "Otros",
}

_LABELS_FUENTE = {
    "pago_membresia": "Pago de membresía",
    "venta_tienda": "Venta de tienda",
    "manual": "Registro manual",
    "pago_directo": "Pago directo (legacy)",
}


@router.get("/exportar-excel")
def exportar_excel(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    q: Optional[str] = Query(None, max_length=100),
    categoria: Optional[str] = Query(None),
    plan_id: Optional[int] = Query(None),
    metodo_pago: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(_require_admin),
):
    """Exporta el módulo financiero del período a .xlsx: Resumen + Ingresos + Egresos.

    Reusa `_recolectar_movimientos` y `balance()`, así que el archivo dice
    exactamente lo mismo que la pantalla. A diferencia del listado, acá no hay
    tope de filas: un export recortado sería un export equivocado.

    **Recibe los mismos filtros que el listado.** Filtrar por "Nómina" y que el archivo
    salga con el período completo es justo el tipo de sorpresa que hace desconfiar del
    export. El Resumen, en cambio, sigue siendo del período sin filtrar: son los totales
    del negocio, y recortarlos por un filtro de pantalla los volvería engañosos.

    La separación por tipo se hace **en memoria, sobre una sola recolección**. Pedirle
    al helper `tipo="ingreso"` y después `tipo="egreso"` sería más corto de escribir,
    pero correría dos veces las consultas de pagos, ventas y movimientos.
    """
    import io

    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    datos = balance(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, db=db, current_user=_)

    # Una sola recolección, particionada en una pasada.
    ingresos, egresos = [], []
    for m in _recolectar_movimientos(
        db, fecha_desde, fecha_hasta,
        q=q, categoria=categoria, plan_id=plan_id, metodo_pago=metodo_pago,
    ):
        (ingresos if m["tipo"] == "ingreso" else egresos).append(m)

    encabezado_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

    def _formatear(ws, filas_encabezado=1):
        for celda in ws[filas_encabezado]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal="center")
        ws.freeze_panes = f"A{filas_encabezado + 1}"
        for idx, col in enumerate(ws.columns, start=1):
            ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(ancho + 2, 12), 50)

    wb = Workbook()

    # ── Hoja 1: Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Concepto", "Monto"])
    periodo = f"{fecha_desde.isoformat() if fecha_desde else 'inicio'} → {fecha_hasta.isoformat() if fecha_hasta else 'hoy'}"
    ws.append(["Período", periodo])
    ws.append(["Membresías", datos.total_membresias])
    ws.append(["Tienda", datos.total_tienda])
    ws.append(["Total ingresos", datos.ingresos_total])
    ws.append(["Egresos", datos.egresos_total])
    ws.append(["Balance neto", datos.balance_neto])
    # El desglose va siempre, aunque no haya egresos: si la sección desaparece
    # cuando está vacía, quien abre el archivo no sabe si no hubo gastos o si el
    # export se rompió.
    ws.append([])
    ws.append(["Egresos por categoría", ""])
    if datos.egresos_por_categoria:
        for cat, val in datos.egresos_por_categoria.items():
            ws.append([_LABELS_CATEGORIA.get(cat, cat), val])
    else:
        ws.append(["Sin egresos en el período", 0])
    _formatear(ws)

    # ── Hojas 2 y 3: Ingresos y Egresos ──
    def _hoja_movimientos(titulo: str, filas: List[dict], con_origen: bool) -> None:
        """Una hoja por tipo. Los montos van positivos: el nombre de la hoja ya dice
        el signo, y así el total de cada hoja se lee directo."""
        hoja = wb.create_sheet(titulo)
        columnas = ["Fecha", "Concepto", "Categoría", "Método de pago", "Cliente"]
        if con_origen:
            columnas.append("Origen")
        columnas.append("Monto")
        hoja.append(columnas)

        for m in filas:
            fecha = m["fecha"]
            fila = [
                # Las fechas se guardan naive en UTC: se pasan a Bogotá para que el
                # Excel coincida con lo que muestra la pantalla.
                fecha.replace(tzinfo=ZoneInfo("UTC")).astimezone(TZ_BOGOTA).strftime("%Y-%m-%d %H:%M") if fecha else None,
                m["concepto"],
                _LABELS_CATEGORIA.get(m["categoria"], m["categoria"]),
                (m["metodo_pago"] or "").capitalize() or None,
                m["usuario_nombre"],
            ]
            if con_origen:
                fila.append(_LABELS_FUENTE.get(m["fuente"], m["fuente"]))
            fila.append(m["monto"])
            hoja.append(fila)

        # Fila de total al pie, para no tener que seleccionar la columna a mano.
        if filas:
            total = [None] * (len(columnas) - 2) + ["TOTAL", sum(m["monto"] for m in filas)]
            hoja.append([])
            hoja.append(total)
            for celda in hoja[hoja.max_row]:
                celda.font = Font(bold=True)
        else:
            hoja.append([f"Sin {titulo.lower()} en el período"])

        _formatear(hoja)

    # "Origen" solo en Ingresos: los egresos son siempre registros manuales, así que
    # la columna diría lo mismo en todas las filas.
    _hoja_movimientos("Ingresos", ingresos, con_origen=True)
    _hoja_movimientos("Egresos", egresos, con_origen=False)

    buffer = io.BytesIO()
    wb.save(buffer)

    nombre_archivo = f"finanzas_jainsportbox_{hoy_bogota().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


@router.post("/movimientos", status_code=status.HTTP_201_CREATED)
def crear_movimiento(
    payload: MovimientoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin),
):
    if payload.usuario_id:
        if not db.query(Usuario).filter(Usuario.id == payload.usuario_id).first():
            raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    mov = MovimientoFinanciero(
        tipo=TipoMovimiento(payload.tipo),
        concepto=payload.concepto,
        categoria=payload.categoria,
        monto=payload.monto,
        fecha=payload.fecha,
        metodo_pago=payload.metodo_pago,
        usuario_id=payload.usuario_id,
        notas=payload.notas,
        fuente="manual",
        created_by=current_user.id,
    )
    db.add(mov)
    db.commit()
    db.refresh(mov)
    return {"id": f"mov_{mov.id}", "mensaje": "Movimiento registrado correctamente."}


@router.delete("/movimientos/{movimiento_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_movimiento(
    movimiento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin),
):
    mov = db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.id == movimiento_id,
        MovimientoFinanciero.fuente == "manual",
    ).first()
    if not mov:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado o no es eliminable.")
    db.delete(mov)
    db.commit()


