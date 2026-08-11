import hmac
import os
from datetime import date, datetime, timedelta
from typing import List, Optional
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import func
from sqlalchemy.orm import Session, defer

from database import get_db
from fechas import FechaInicio, hoy_bogota
from membresia import aplicar_personalizado, aplicar_plan
from models import MovimientoFinanciero, Pago, Plan, RolUsuario, TipoMovimiento, Usuario
from schemas.usuario import UsuarioCreate, UsuarioResponse, UsuarioUpdate
from security import get_current_user, get_password_hash
from storage import guardar_archivo, eliminar_archivo

ALLOWED_TYPES = {"image/jpeg", "image/png", "image/webp"}

router = APIRouter(prefix="/usuarios", tags=["Usuarios"])


def _require_admin_or_coach(current_user: Usuario = Depends(get_current_user)):
    if current_user.rol not in (RolUsuario.ADMIN, RolUsuario.COACH):
        raise HTTPException(status_code=403, detail="Solo admin o coach pueden realizar esta acción.")
    return current_user


def _require_admin(current_user: Usuario = Depends(get_current_user)):
    if current_user.rol != RolUsuario.ADMIN:
        raise HTTPException(status_code=403, detail="Solo un admin puede realizar esta acción.")
    return current_user


# Roles con privilegio de staff: solo un admin puede crear/tocar estas cuentas.
_ROLES_PRIVILEGIADOS = (RolUsuario.ADMIN, RolUsuario.COACH)


@router.post("/", response_model=UsuarioResponse, status_code=status.HTTP_201_CREATED)
def crear_usuario(
    payload: UsuarioCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    # El admin es único y lo siembra seed.py: no se crea otro por API, ni siquiera
    # siendo admin. Sin esto, cualquier admin podría fabricarse un segundo dueño del
    # box (y el guard de auto-borrado dejaría de garantizar que siempre haya uno solo).
    if payload.rol == RolUsuario.ADMIN:
        raise HTTPException(status_code=403, detail="No se pueden crear cuentas de administrador.")

    # Un coach no puede crear cuentas de staff; solo el admin.
    if payload.rol in _ROLES_PRIVILEGIADOS and current_user.rol != RolUsuario.ADMIN:
        raise HTTPException(status_code=403, detail="Solo el administrador puede crear cuentas de coach.")

    email_norm = (payload.email or "").strip().lower()
    doc_norm = (payload.documento_identidad or "").strip()
    if db.query(Usuario).filter(Usuario.email == email_norm).first():
        raise HTTPException(status_code=400, detail="Ya existe un usuario con ese email.")
    if db.query(Usuario).filter(Usuario.documento_identidad == doc_norm).first():
        raise HTTPException(status_code=400, detail="Ya existe un usuario con ese documento de identidad.")
    nuevo = Usuario(
        nombre=payload.nombre,
        email=email_norm,
        password_hash=get_password_hash(payload.password),
        documento_identidad=doc_norm,
        genero=payload.genero,
        rol=payload.rol,
        huella_id=payload.huella_id,
        telefono=payload.telefono,
        fecha_nacimiento=payload.fecha_nacimiento,
        eps=payload.eps,
        barrio=payload.barrio,
        contacto_emergencia_nombre=payload.contacto_emergencia_nombre,
        contacto_emergencia_telefono=payload.contacto_emergencia_telefono,
        es_menor=payload.es_menor,
        acudiente_nombre=payload.acudiente_nombre,
        acudiente_telefono=payload.acudiente_telefono,
        acudiente_documento=payload.acudiente_documento,
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo


@router.patch("/{usuario_id}", response_model=UsuarioResponse)
def actualizar_usuario(
    usuario_id: int,
    payload: UsuarioUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    # Un coach no puede modificar (ni resetear la contraseña de) una cuenta de staff.
    if usuario.rol in _ROLES_PRIVILEGIADOS and current_user.rol != RolUsuario.ADMIN:
        raise HTTPException(status_code=403, detail="Solo el administrador puede modificar cuentas de admin o coach.")

    if payload.nombre is not None:
        usuario.nombre = payload.nombre

    if payload.email is not None:
        email_norm = payload.email.strip().lower()
        duplicado = (
            db.query(Usuario)
            .filter(Usuario.email == email_norm, Usuario.id != usuario_id)
            .first()
        )
        if duplicado:
            raise HTTPException(status_code=400, detail="Ya existe un usuario con ese email.")
        usuario.email = email_norm

    if payload.password is not None:
        usuario.password_hash = get_password_hash(payload.password)

    if payload.telefono is not None:
        usuario.telefono = payload.telefono

    if payload.documento_identidad is not None:
        doc_norm = payload.documento_identidad.strip()
        duplicado_doc = (
            db.query(Usuario)
            .filter(Usuario.documento_identidad == doc_norm, Usuario.id != usuario_id)
            .first()
        )
        if duplicado_doc:
            raise HTTPException(status_code=400, detail="Ya existe un usuario con ese documento de identidad.")
        usuario.documento_identidad = doc_norm

    if payload.genero is not None:
        usuario.genero = payload.genero

    if payload.fecha_nacimiento is not None:
        usuario.fecha_nacimiento = payload.fecha_nacimiento

    if payload.eps is not None:
        usuario.eps = payload.eps

    if payload.barrio is not None:
        usuario.barrio = payload.barrio

    if payload.contacto_emergencia_nombre is not None:
        usuario.contacto_emergencia_nombre = payload.contacto_emergencia_nombre

    if payload.contacto_emergencia_telefono is not None:
        usuario.contacto_emergencia_telefono = payload.contacto_emergencia_telefono

    if payload.es_menor is not None:
        usuario.es_menor = payload.es_menor

    if payload.acudiente_nombre is not None:
        usuario.acudiente_nombre = payload.acudiente_nombre

    if payload.acudiente_telefono is not None:
        usuario.acudiente_telefono = payload.acudiente_telefono

    if payload.acudiente_documento is not None:
        usuario.acudiente_documento = payload.acudiente_documento

    db.commit()
    db.refresh(usuario)
    return usuario


@router.post("/{usuario_id}/foto", response_model=UsuarioResponse)
def subir_foto(
    usuario_id: int,
    foto: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    if foto.content_type not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Formato no permitido. Usa JPG, PNG o WEBP.")

    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    if usuario.foto_url:
        eliminar_archivo(usuario.foto_url)

    usuario.foto_url = guardar_archivo(foto)
    db.commit()
    db.refresh(usuario)
    return usuario


@router.delete("/{usuario_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_usuario(
    usuario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
    # Va primero que el guard de rol: es absoluto y aplica también al admin, que si
    # no podría dejar el sistema sin ninguna cuenta de administración.
    if usuario.id == current_user.id:
        raise HTTPException(status_code=400, detail="No podés eliminar tu propia cuenta.")
    # Mismo criterio que crear/editar: las cuentas de staff solo las toca un admin.
    if usuario.rol in _ROLES_PRIVILEGIADOS and current_user.rol != RolUsuario.ADMIN:
        raise HTTPException(
            status_code=403,
            detail="Solo un admin puede eliminar cuentas de staff.",
        )
    if usuario.foto_url:
        eliminar_archivo(usuario.foto_url)
    db.delete(usuario)
    db.commit()


@router.get("/", response_model=List[UsuarioResponse])
def listar_usuarios(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    # defer: no traer la columna Text pesada (huella_template) en el listado.
    return (
        db.query(Usuario)
        .options(defer(Usuario.huella_template))
        .filter(Usuario.rol != RolUsuario.PENDIENTE)
        .all()
    )


# IMPORTANTE: rutas estáticas ANTES de las parametrizadas con {usuario_id}.
# FastAPI matchea en orden de declaración; si /{usuario_id} se declara primero,
# captura "pendientes" como ID y falla con 422.
@router.get("/pendientes")
def listar_pendientes(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    pendientes = db.query(Usuario).filter(Usuario.rol == RolUsuario.PENDIENTE).all()
    result = []
    for u in pendientes:
        result.append({
            "id": u.id,
            "nombre": u.nombre,
            "email": u.email,
            "telefono": u.telefono,
            "documento_identidad": u.documento_identidad,
            "foto_url": u.foto_url,
            "genero": u.genero,
            "fecha_nacimiento": u.fecha_nacimiento,
            "eps": u.eps,
            "barrio": u.barrio,
            "contacto_emergencia_nombre": u.contacto_emergencia_nombre,
            "contacto_emergencia_telefono": u.contacto_emergencia_telefono,
            # Menor de edad: el admin debe ver los datos del acudiente antes de activar.
            "es_menor": u.es_menor,
            "acudiente_nombre": u.acudiente_nombre,
            "acudiente_telefono": u.acudiente_telefono,
            "acudiente_documento": u.acudiente_documento,
            "created_at": u.created_at,
        })
    return result


class EliminarPendientesRequest(BaseModel):
    ids: List[int] = Field(..., min_length=1, max_length=500)


@router.post("/pendientes/eliminar")
def eliminar_pendientes(
    payload: EliminarPendientesRequest,
    db: Session = Depends(get_db),
    _: Usuario = Depends(_require_admin),
):
    """Descarta en bloque solicitudes de registro que nunca se activaron.

    El filtro por `rol == PENDIENTE` es la propiedad de seguridad del endpoint: si
    entre los ids llega el de un cliente activo o el de un coach, se ignora en vez
    de borrarse. Así un id equivocado (o manipulado) no puede tumbar una cuenta real.
    """
    pendientes = (
        db.query(Usuario)
        .filter(Usuario.id.in_(payload.ids), Usuario.rol == RolUsuario.PENDIENTE)
        .all()
    )
    for u in pendientes:
        if u.foto_url:
            eliminar_archivo(u.foto_url)
        db.delete(u)
    db.commit()
    return {"eliminados": len(pendientes)}


@router.get("/exportar-excel")
def exportar_excel(
    clientes: bool = Query(True, description="Incluir la hoja de clientes."),
    equipo: bool = Query(False, description="Incluir la hoja del equipo del box (admin y coaches)."),
    db: Session = Depends(get_db),
    _: Usuario = Depends(_require_admin_or_coach),
):
    """Exporta a .xlsx los grupos pedidos, una hoja por grupo.

    Las dos poblaciones no comparten columnas: al staff no le aplican membresía,
    acudiente ni términos, así que cada hoja lleva su propio encabezado en vez de
    una tabla común llena de celdas vacías.
    """
    import io

    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not clientes and not equipo:
        raise HTTPException(status_code=400, detail="Selecciona al menos un grupo para exportar.")

    hoy = hoy_bogota()

    def _edad(fn):
        if not fn:
            return None
        return hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))

    def _consultar(*filtros):
        return (
            db.query(Usuario)
            .options(defer(Usuario.huella_template))
            .filter(*filtros)
            .order_by(Usuario.nombre.asc())
            .all()
        )

    encabezado_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

    def _formatear(ws):
        """Encabezado rojo, panel congelado y ancho de columna según contenido."""
        for celda in ws[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = encabezado_fill
            celda.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for idx, col in enumerate(ws.columns, start=1):
            ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(ancho + 2, 10), 40)

    wb = Workbook()
    wb.remove(wb.active)  # las hojas se crean según lo pedido

    if clientes:
        ws = wb.create_sheet("Clientes")
        ws.append([
            "Nombre", "Documento", "Email", "Teléfono", "Género",
            "Fecha nacimiento", "Edad", "EPS", "Barrio",
            "Emergencia: nombre", "Emergencia: teléfono",
            "Menor de edad", "Acudiente: nombre", "Acudiente: cédula", "Acudiente: teléfono",
            "Membresía", "Vence", "Días restantes",
            "Huella", "Términos aceptados", "Fecha aceptación", "Versión términos",
            "Registrado",
        ])
        for u in _consultar(Usuario.rol == RolUsuario.CLIENTE):
            if u.fecha_vencimiento:
                dias = (u.fecha_vencimiento - hoy).days
                membresia = "Activa" if dias >= 0 else "Vencida"
            else:
                dias = None
                membresia = "Sin plan"
            ws.append([
                u.nombre,
                u.documento_identidad,
                u.email,
                u.telefono,
                u.genero,
                u.fecha_nacimiento.isoformat() if u.fecha_nacimiento else None,
                _edad(u.fecha_nacimiento),
                u.eps,
                u.barrio,
                u.contacto_emergencia_nombre,
                u.contacto_emergencia_telefono,
                "Sí" if u.es_menor else "No",
                u.acudiente_nombre,
                u.acudiente_documento,
                u.acudiente_telefono,
                membresia,
                u.fecha_vencimiento.isoformat() if u.fecha_vencimiento else None,
                dias,
                "Registrada" if u.huella_id else "No",
                "Sí" if u.acepto_terminos else "No",
                u.terminos_fecha.strftime("%Y-%m-%d %H:%M") if u.terminos_fecha else None,
                u.terminos_version,
                u.created_at.strftime("%Y-%m-%d") if u.created_at else None,
            ])
        _formatear(ws)

    if equipo:
        ws = wb.create_sheet("Equipo del box")
        ws.append([
            "Nombre", "Rol", "Documento", "Email", "Teléfono", "Género",
            "Fecha nacimiento", "Edad", "EPS", "Barrio",
            "Emergencia: nombre", "Emergencia: teléfono",
            "Huella", "Registrado",
        ])
        for u in _consultar(Usuario.rol.in_(_ROLES_PRIVILEGIADOS)):
            ws.append([
                u.nombre,
                u.rol.value,
                u.documento_identidad,
                u.email,
                u.telefono,
                u.genero,
                u.fecha_nacimiento.isoformat() if u.fecha_nacimiento else None,
                _edad(u.fecha_nacimiento),
                u.eps,
                u.barrio,
                u.contacto_emergencia_nombre,
                u.contacto_emergencia_telefono,
                "Registrada" if u.huella_id else "No",
                u.created_at.strftime("%Y-%m-%d") if u.created_at else None,
            ])
        _formatear(ws)

    buffer = io.BytesIO()
    wb.save(buffer)

    grupo = "clientes" if clientes and not equipo else "equipo" if equipo and not clientes else "usuarios"
    nombre_archivo = f"{grupo}_jainsportbox_{hoy.isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


def query_cumpleaneros_hoy(db: Session) -> List[Usuario]:
    """Socios con membresía vigente que cumplen años hoy (hora Bogotá).

    Vive acá y lo consume también el router de dashboard, para que el criterio
    (incluida la exclusión de membresías vencidas) no se duplique.
    """
    hoy = hoy_bogota()
    # extract() es portable (SQLite y Postgres); strftime es solo de SQLite.
    return (
        db.query(Usuario)
        .filter(
            Usuario.fecha_nacimiento.isnot(None),
            func.extract("month", Usuario.fecha_nacimiento) == hoy.month,
            func.extract("day", Usuario.fecha_nacimiento) == hoy.day,
            Usuario.fecha_vencimiento >= hoy,
        )
        .all()
    )


@router.get("/{usuario_id}", response_model=UsuarioResponse)
def obtener_usuario(
    usuario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
    return usuario


class ActivarUsuarioPayload(BaseModel):
    """Activación de un pendiente, con plan del catálogo o membresía personalizada.

    `plan_id` y `duracion_dias` son excluyentes: uno define la membresía por el
    catálogo y el otro a mano. Aceptar los dos obligaría a elegir cuál gana, y
    cualquier criterio sería una sorpresa para quien cargó el otro.
    """
    plan_id: Optional[int] = None
    duracion_dias: Optional[int] = Field(None, ge=1, le=365)
    numero_ingresos: Optional[int] = Field(None, ge=1)
    monto: float = Field(..., ge=0)
    metodo_pago: str = Field(..., pattern=r'^(efectivo|transferencia)$')
    fecha_inicio: FechaInicio = None

    @model_validator(mode="after")
    def _plan_o_dias(self):
        if bool(self.plan_id) == bool(self.duracion_dias):
            raise ValueError("Indicá un plan del catálogo o una cantidad de días, pero no ambos.")
        return self


@router.post("/{usuario_id}/activar")
def activar_usuario(
    usuario_id: int,
    payload: ActivarUsuarioPayload,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(_require_admin_or_coach),
):
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id, Usuario.rol == RolUsuario.PENDIENTE).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario pendiente no encontrado.")

    plan = None
    if payload.plan_id:
        plan = db.query(Plan).filter(Plan.id == payload.plan_id, Plan.activo == True).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Plan no encontrado.")

    usuario.rol = RolUsuario.CLIENTE
    # Las dos funciones cargan también los accesos. Un pendiente no tiene vencimiento
    # previo, así que la base termina siendo hoy (o la fecha de inicio, si es futura).
    hoy = hoy_bogota()
    if plan:
        nueva_fecha = aplicar_plan(usuario, plan, hoy, payload.fecha_inicio)
    else:
        nueva_fecha = aplicar_personalizado(
            usuario, payload.duracion_dias, payload.numero_ingresos, hoy, payload.fecha_inicio
        )

    pago = Pago(
        usuario_id=usuario.id,
        plan_id=plan.id if plan else None,
        duracion_dias=None if plan else payload.duracion_dias,
        numero_ingresos=None if plan else payload.numero_ingresos,
        monto=payload.monto,
        metodo_pago=payload.metodo_pago,
        fecha_inicio=payload.fecha_inicio,
    )
    db.add(pago)

    db.commit()
    return {
        "message": f"Usuario {usuario.nombre} activado correctamente.",
        "usuario_id": usuario.id,
        "nueva_fecha_vencimiento": nueva_fecha,
    }


class HuellaTemplatePayload(BaseModel):
    template: str  # base64-encoded FMD


@router.post("/{usuario_id}/huella-template", status_code=status.HTTP_200_OK)
def guardar_huella_template(
    usuario_id: int,
    payload: HuellaTemplatePayload,
    request: Request,
    db: Session = Depends(get_db),
):
    """Guarda el template de huella. Acepta JWT de admin/coach O el header X-Bridge-Secret del bridge .NET."""
    bridge_secret = os.environ.get("BRIDGE_SECRET", "")
    x_secret = request.headers.get("X-Bridge-Secret", "")
    if not bridge_secret or not hmac.compare_digest(x_secret, bridge_secret):
        # Si no viene del bridge, exigir JWT admin/coach
        try:
            from fastapi.security import OAuth2PasswordBearer
            from security import get_current_user
            token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
            if not token:
                raise HTTPException(status_code=401, detail="Not authenticated")
            from security import SECRET_KEY, ALGORITHM
            from jose import jwt as jose_jwt, JWTError
            payload_jwt = jose_jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            email = payload_jwt.get("sub")
            caller = db.query(Usuario).filter(Usuario.email == email).first()
            if not caller or caller.rol.value not in ("admin", "coach"):
                raise HTTPException(status_code=403, detail="Sin permisos.")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=401, detail="Not authenticated")

    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
    usuario.huella_template = payload.template
    usuario.huella_id = f"dp_{usuario_id}"
    db.commit()
    return {"mensaje": f"Huella registrada para {usuario.nombre}."}


@router.get("/con-template/lista")
def listar_usuarios_con_template(
    request: Request,
    db: Session = Depends(get_db),
):
    """Devuelve id, nombre y template de todos los usuarios con huella. Acepta JWT admin/coach o X-Bridge-Secret."""
    bridge_secret = os.environ.get("BRIDGE_SECRET", "")
    x_secret = request.headers.get("X-Bridge-Secret", "")
    if not bridge_secret or not hmac.compare_digest(x_secret, bridge_secret):
        try:
            from jose import jwt as jose_jwt, JWTError
            from security import SECRET_KEY, ALGORITHM
            token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
            if not token:
                raise HTTPException(status_code=401, detail="Not authenticated")
            payload_jwt = jose_jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            email = payload_jwt.get("sub")
            caller = db.query(Usuario).filter(Usuario.email == email).first()
            if not caller or caller.rol.value not in ("admin", "coach"):
                raise HTTPException(status_code=403, detail="Sin permisos.")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=401, detail="Not authenticated")
    usuarios = (
        db.query(Usuario.id, Usuario.nombre, Usuario.huella_template)
        .filter(Usuario.huella_template.isnot(None))
        .all()
    )
    return [
        {"id": u.id, "nombre": u.nombre, "template": u.huella_template}
        for u in usuarios
    ]
