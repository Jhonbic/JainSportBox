"""§3 de tests.md — Usuarios: CRUD, roles, cumpleaños, huella/bridge."""

import uuid
from datetime import date, timedelta

import models
from conftest import PNG_BYTES, SessionLocal, auth_headers
from fechas import hoy_bogota


def _payload_usuario(**overrides):
    sufijo = uuid.uuid4().hex[:8]
    data = {
        "nombre": f"Nuevo {sufijo}",
        "email": f"nuevo{sufijo}@test.local",
        "password": "claveValida1",
        "documento_identidad": f"c{sufijo}",
        "genero": "femenino",
        "rol": "cliente",
        "telefono": "3011112233",
        "contacto_emergencia_nombre": "Contacto Emergencia",
        "contacto_emergencia_telefono": "3111112233",
    }
    data.update(overrides)
    return data


# ── POST /usuarios/ ────────────────────────────────────────────


def test_admin_crea_cliente(client, admin_headers):
    r = client.post("/usuarios/", json=_payload_usuario(), headers=admin_headers)
    assert r.status_code == 201
    body = r.json()
    assert body["rol"] == "cliente"
    # la respuesta nunca expone credenciales ni datos biométricos
    assert "password" not in body and "password_hash" not in body
    assert "huella_template" not in body


def test_contacto_de_emergencia_es_obligatorio(client, admin_headers):
    """Mismo criterio que el registro público: sin contacto de emergencia no se crea."""
    payload = _payload_usuario()
    del payload["contacto_emergencia_nombre"]
    del payload["contacto_emergencia_telefono"]
    assert client.post("/usuarios/", json=payload, headers=admin_headers).status_code == 422


def test_coach_crea_cliente_pero_no_staff(client, coach):
    assert client.post("/usuarios/", json=_payload_usuario(), headers=coach.headers).status_code == 201
    assert client.post("/usuarios/", json=_payload_usuario(rol="coach"), headers=coach.headers).status_code == 403
    assert client.post("/usuarios/", json=_payload_usuario(rol="admin"), headers=coach.headers).status_code == 403


def test_admin_crea_coach(client, admin_headers):
    r = client.post("/usuarios/", json=_payload_usuario(rol="coach"), headers=admin_headers)
    assert r.status_code == 201


def test_nadie_puede_crear_otro_admin(client, admin_headers):
    """El admin es único (lo siembra seed.py): ni el propio admin crea otro."""
    r = client.post("/usuarios/", json=_payload_usuario(rol="admin"), headers=admin_headers)
    assert r.status_code == 403


def test_cliente_no_crea_usuarios(client, cliente):
    assert client.post("/usuarios/", json=_payload_usuario(), headers=cliente.headers).status_code == 403


def test_crear_sin_token(client):
    assert client.post("/usuarios/", json=_payload_usuario()).status_code == 401


def test_email_duplicado_case_insensitive(client, admin_headers, cliente):
    r = client.post(
        "/usuarios/",
        json=_payload_usuario(email=cliente.user.email.upper()),
        headers=admin_headers,
    )
    assert r.status_code == 400


# ── GET /usuarios/ ─────────────────────────────────────────────


def test_listado_excluye_pendientes_y_template(client, admin_headers, cliente, pendiente):
    r = client.get("/usuarios/", headers=admin_headers)
    assert r.status_code == 200
    body = r.json()
    ids = [u["id"] for u in body]
    assert cliente.user.id in ids
    assert pendiente.user.id not in ids
    assert all("huella_template" not in u for u in body)


def test_listado_cliente_403(client, cliente):
    assert client.get("/usuarios/", headers=cliente.headers).status_code == 403


def test_pendientes(client, admin_headers, pendiente, cliente, db_session):
    db_session.query(models.Usuario).filter_by(id=pendiente.user.id).update(
        {"foto_url": "fotos/pendiente.jpg"}
    )
    db_session.commit()
    r = client.get("/usuarios/pendientes", headers=admin_headers)
    assert r.status_code == 200
    body = r.json()
    assert [u["id"] for u in body] == [pendiente.user.id]
    # El listado muestra el avatar del pendiente, así que la foto viaja en el payload.
    assert body[0]["foto_url"] == "fotos/pendiente.jpg"


# ── GET /usuarios/exportar-excel ───────────────────────────────


def test_exportar_excel_admin(client, admin_headers, cliente):
    r = client.get("/usuarios/exportar-excel", headers=admin_headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "clientes_jainsportbox_" in r.headers["content-disposition"]

    # El .xlsx es válido y contiene al cliente con sus columnas
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    encabezados = filas[0]
    assert "Nombre" in encabezados and "EPS" in encabezados and "Acudiente: cédula" in encabezados
    nombres = [f[0] for f in filas[1:]]
    assert cliente.user.nombre in nombres


def test_exportar_excel_coach_ok(client, coach):
    assert client.get("/usuarios/exportar-excel", headers=coach.headers).status_code == 200


def test_exportar_excel_solo_equipo(client, admin_headers, coach, cliente):
    """equipo=true trae la hoja del staff, y el cliente no se cuela ahí."""
    import io

    from openpyxl import load_workbook

    r = client.get(
        "/usuarios/exportar-excel",
        params={"clientes": "false", "equipo": "true"},
        headers=admin_headers,
    )
    assert r.status_code == 200
    assert "equipo_jainsportbox_" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Equipo del box"]
    filas = list(wb["Equipo del box"].iter_rows(values_only=True))
    assert "Rol" in filas[0]
    nombres = [f[0] for f in filas[1:]]
    assert coach.user.nombre in nombres
    assert cliente.user.nombre not in nombres


def test_exportar_excel_ambos_grupos(client, admin_headers, coach, cliente):
    import io

    from openpyxl import load_workbook

    r = client.get(
        "/usuarios/exportar-excel",
        params={"clientes": "true", "equipo": "true"},
        headers=admin_headers,
    )
    assert r.status_code == 200
    assert "usuarios_jainsportbox_" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Clientes", "Equipo del box"]


def test_exportar_excel_sin_grupos_400(client, admin_headers):
    r = client.get(
        "/usuarios/exportar-excel",
        params={"clientes": "false", "equipo": "false"},
        headers=admin_headers,
    )
    assert r.status_code == 400


def test_exportar_excel_cliente_403(client, cliente):
    assert client.get("/usuarios/exportar-excel", headers=cliente.headers).status_code == 403


# ── GET /usuarios/{id} ─────────────────────────────────────────


def test_obtener_usuario(client, admin_headers, cliente):
    r = client.get(f"/usuarios/{cliente.user.id}", headers=admin_headers)
    assert r.status_code == 200
    assert r.json()["email"] == cliente.user.email


def test_obtener_usuario_expone_el_arranque_futuro(client, admin_headers, cliente):
    """`membresia_inicio` es la compuerta que niega el acceso antes que la fecha, y sin
    exponerla el perfil mostraba "N días restantes" en verde mientras la palanquera
    rechazaba a la persona: el admin no tenía cómo ver por qué."""
    inicio = hoy_bogota() + timedelta(days=10)
    client.post(
        "/pagos/directo/",
        json={
            "usuario_id": cliente.user.id,
            "duracion_dias": 30,
            "monto": 100,
            "metodo_pago": "efectivo",
            "fecha_inicio": inicio.isoformat(),
        },
        headers=admin_headers,
    )
    body = client.get(f"/usuarios/{cliente.user.id}", headers=admin_headers).json()
    assert body["membresia_inicio"] == inicio.isoformat()


def test_obtener_usuario_sin_arranque_futuro_va_en_null(client, admin_headers, cliente):
    """Una membresía que ya está corriendo no tiene compuerta: el campo va en null y la
    card no muestra el aviso."""
    body = client.get(f"/usuarios/{cliente.user.id}", headers=admin_headers).json()
    assert body["membresia_inicio"] is None


def test_obtener_usuario_id_no_numerico(client, admin_headers):
    assert client.get("/usuarios/abc", headers=admin_headers).status_code == 422


def test_obtener_usuario_inexistente(client, admin_headers):
    assert client.get("/usuarios/999999", headers=admin_headers).status_code == 404


# ── PATCH /usuarios/{id} ───────────────────────────────────────


def test_patch_usuario(client, admin_headers, cliente):
    r = client.patch(
        f"/usuarios/{cliente.user.id}",
        json={"nombre": "Editado", "telefono": "3022223344"},
        headers=admin_headers,
    )
    assert r.status_code == 200
    assert r.json()["nombre"] == "Editado"


def test_coach_no_modifica_staff(client, coach, crear_usuario):
    otro_coach = crear_usuario("coach")
    r = client.patch(
        f"/usuarios/{otro_coach.user.id}", json={"nombre": "Hackeado"}, headers=coach.headers
    )
    assert r.status_code == 403


def test_patch_email_duplicado(client, admin_headers, cliente, coach):
    r = client.patch(
        f"/usuarios/{cliente.user.id}",
        json={"email": coach.user.email},
        headers=admin_headers,
    )
    assert r.status_code == 400


# ── POST /usuarios/{id}/activar ────────────────────────────────


def test_activar_pendiente(client, admin_headers, pendiente, db_session):
    plan = db_session.query(models.Plan).filter_by(nombre="1 Mes").first()
    r = client.post(
        f"/usuarios/{pendiente.user.id}/activar",
        json={"plan_id": plan.id, "monto": 100000, "metodo_pago": "efectivo"},
        headers=admin_headers,
    )
    assert r.status_code == 200
    esperado = (date.today() + timedelta(days=30)).isoformat()
    assert r.json()["nueva_fecha_vencimiento"] == esperado

    u = db_session.query(models.Usuario).filter_by(id=pendiente.user.id).first()
    db_session.refresh(u)
    assert u.rol == models.RolUsuario.CLIENTE
    pago = db_session.query(models.Pago).filter_by(usuario_id=u.id).first()
    assert pago is not None and pago.monto == 100000


def test_activar_no_pendiente_404(client, admin_headers, cliente, db_session):
    plan = db_session.query(models.Plan).first()
    r = client.post(
        f"/usuarios/{cliente.user.id}/activar",
        json={"plan_id": plan.id, "monto": 1000, "metodo_pago": "efectivo"},
        headers=admin_headers,
    )
    assert r.status_code == 404


def test_activar_metodo_invalido_422(client, admin_headers, pendiente, db_session):
    plan = db_session.query(models.Plan).first()
    r = client.post(
        f"/usuarios/{pendiente.user.id}/activar",
        json={"plan_id": plan.id, "monto": 1000, "metodo_pago": "bitcoin"},
        headers=admin_headers,
    )
    assert r.status_code == 422


# ── DELETE /usuarios/{id} ──────────────────────────────────────


def test_eliminar_usuario_cascada(client, admin_headers, cliente, db_session):
    uid = cliente.user.id
    plan = db_session.query(models.Plan).first()
    db_session.add(models.Pago(usuario_id=uid, plan_id=plan.id, monto=1000, metodo_pago="efectivo"))
    db_session.add(models.Asistencia(usuario_id=uid, tipo="entrada"))
    db_session.add(models.MarcaRM(usuario_id=uid, ejercicio="Back Squat", peso=100, repeticiones=1, rm_calculado=100, fecha=date.today()))
    db_session.commit()

    assert client.delete(f"/usuarios/{uid}", headers=admin_headers).status_code == 204

    db = SessionLocal()
    try:
        assert db.query(models.Usuario).filter_by(id=uid).first() is None
        # cascade all, delete-orphan en el modelo → los hijos se van con el usuario
        assert db.query(models.Pago).filter_by(usuario_id=uid).count() == 0
        assert db.query(models.Asistencia).filter_by(usuario_id=uid).count() == 0
        assert db.query(models.MarcaRM).filter_by(usuario_id=uid).count() == 0
    finally:
        db.close()


def test_eliminar_usuario_inexistente(client, admin_headers):
    assert client.delete("/usuarios/999999", headers=admin_headers).status_code == 404


def test_coach_no_puede_eliminar_admin(client, coach, crear_usuario):
    otro_admin = crear_usuario("admin")
    r = client.delete(f"/usuarios/{otro_admin.user.id}", headers=coach.headers)
    assert r.status_code == 403


def test_coach_no_puede_eliminar_otro_coach(client, coach, crear_usuario):
    otro_coach = crear_usuario("coach")
    r = client.delete(f"/usuarios/{otro_coach.user.id}", headers=coach.headers)
    assert r.status_code == 403


def test_coach_si_puede_eliminar_cliente(client, coach, cliente):
    r = client.delete(f"/usuarios/{cliente.user.id}", headers=coach.headers)
    assert r.status_code == 204


# ── POST /usuarios/pendientes/eliminar ─────────────────────────


def test_eliminar_pendientes_en_bloque(client, admin_headers, crear_usuario):
    ids = [crear_usuario("pendiente").user.id for _ in range(3)]
    r = client.post("/usuarios/pendientes/eliminar", json={"ids": ids}, headers=admin_headers)
    assert r.status_code == 200
    assert r.json()["eliminados"] == 3
    assert client.get("/usuarios/pendientes", headers=admin_headers).json() == []


def test_eliminar_pendientes_ignora_a_los_que_no_son_pendientes(
    client, admin_headers, cliente, coach, pendiente
):
    # El filtro por rol es lo que impide que un id equivocado borre una cuenta real.
    r = client.post(
        "/usuarios/pendientes/eliminar",
        json={"ids": [pendiente.user.id, cliente.user.id, coach.user.id]},
        headers=admin_headers,
    )
    assert r.status_code == 200
    assert r.json()["eliminados"] == 1
    ids_vivos = [u["id"] for u in client.get("/usuarios/", headers=admin_headers).json()]
    assert cliente.user.id in ids_vivos
    assert coach.user.id in ids_vivos


def test_eliminar_pendientes_coach_403(client, coach, pendiente):
    r = client.post(
        "/usuarios/pendientes/eliminar",
        json={"ids": [pendiente.user.id]},
        headers=coach.headers,
    )
    assert r.status_code == 403


def test_staff_puede_eliminar_pendiente(client, coach, pendiente):
    # Descartar la solicitud del que se registró y nunca apareció por el box.
    r = client.delete(f"/usuarios/{pendiente.user.id}", headers=coach.headers)
    assert r.status_code == 204
    assert client.get("/usuarios/pendientes", headers=coach.headers).json() == []


def test_admin_puede_eliminar_coach(client, admin_headers, coach):
    r = client.delete(f"/usuarios/{coach.user.id}", headers=admin_headers)
    assert r.status_code == 204


def test_nadie_puede_eliminarse_a_si_mismo(client, coach):
    # El guard aplica también al admin: si no, podría dejar el sistema sin administración.
    assert client.delete(f"/usuarios/{coach.user.id}", headers=coach.headers).status_code == 400


# ── Foto de perfil (admin) ─────────────────────────────────────


def test_subir_foto_usuario(client, admin_headers, cliente):
    r = client.post(
        f"/usuarios/{cliente.user.id}/foto",
        files={"foto": ("p.png", PNG_BYTES, "image/png")},
        headers=admin_headers,
    )
    assert r.status_code == 200
    assert r.json()["foto_url"].startswith("/uploads/")


# ── Huella: bridge vs JWT ──────────────────────────────────────


def test_guardar_template_con_bridge_secret(client, bridge_headers, cliente, db_session):
    r = client.post(
        f"/usuarios/{cliente.user.id}/huella-template",
        json={"template": "dGVtcGxhdGUtZmFrZQ=="},
        headers=bridge_headers,
    )
    assert r.status_code == 200
    u = db_session.query(models.Usuario).filter_by(id=cliente.user.id).first()
    assert u.huella_id == f"dp_{cliente.user.id}"
    assert u.huella_template == "dGVtcGxhdGUtZmFrZQ=="


def test_guardar_template_secret_incorrecto(client, cliente):
    r = client.post(
        f"/usuarios/{cliente.user.id}/huella-template",
        json={"template": "x"},
        headers={"X-Bridge-Secret": "incorrecto"},
    )
    assert r.status_code == 401


def test_guardar_template_jwt_cliente_403(client, cliente):
    r = client.post(
        f"/usuarios/{cliente.user.id}/huella-template",
        json={"template": "x"},
        headers=cliente.headers,
    )
    assert r.status_code == 403


def test_guardar_template_jwt_admin_ok(client, admin_headers, cliente):
    r = client.post(
        f"/usuarios/{cliente.user.id}/huella-template",
        json={"template": "abc"},
        headers=admin_headers,
    )
    assert r.status_code == 200


def test_lista_con_template(client, bridge_headers, cliente, coach, db_session):
    db_session.query(models.Usuario).filter_by(id=cliente.user.id).update(
        {"huella_template": "t1", "huella_id": f"dp_{cliente.user.id}"}
    )
    db_session.commit()
    r = client.get("/usuarios/con-template/lista", headers=bridge_headers)
    assert r.status_code == 200
    body = r.json()
    assert [u["id"] for u in body] == [cliente.user.id]
    assert body[0]["template"] == "t1"


def test_lista_con_template_sin_auth(client):
    assert client.get("/usuarios/con-template/lista").status_code == 401


