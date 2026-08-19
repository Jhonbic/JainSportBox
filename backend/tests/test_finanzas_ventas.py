"""§9 de tests.md — Finanzas (balance sin doble conteo, movimientos), Ventas
(stock) y Productos (CRUD, soft delete)."""

from datetime import date, datetime, timedelta

import models


def _crear_producto(client, admin_headers, precio=20000, stock=10, **kw):
    payload = {"nombre": kw.pop("nombre", "Batido Test"), "precio": precio, "stock": stock}
    payload.update(kw)
    r = client.post("/productos/", json=payload, headers=admin_headers)
    assert r.status_code == 201
    return r.json()


def _vender(client, headers, producto_id, cantidad=1):
    return client.post(
        "/ventas/",
        json={"producto_id": producto_id, "cantidad": cantidad, "metodo_pago": "efectivo"},
        headers=headers,
    )


def _mov(client, admin_headers, tipo, monto, categoria="ingreso_varios", concepto=None):
    return client.post(
        "/finanzas/movimientos",
        json={
            "tipo": tipo,
            "concepto": concepto or f"{tipo} test",
            "categoria": categoria,
            "monto": monto,
            "fecha": datetime.utcnow().isoformat(),
        },
        headers=admin_headers,
    )


def _pagar(client, admin_headers, usuario_id, plan_id, monto=35000, metodo="efectivo"):
    r = client.post(
        "/pagos/",
        json={"usuario_id": usuario_id, "plan_id": plan_id, "monto": monto, "metodo_pago": metodo},
        headers=admin_headers,
    )
    assert r.status_code == 201
    return r.json()


# ── Productos ──────────────────────────────────────────────────


def test_producto_crud_y_soft_delete(client, admin_headers, cliente):
    prod = _crear_producto(client, admin_headers)

    # editar
    r = client.put(f"/productos/{prod['id']}", json={"precio": 25000}, headers=admin_headers)
    assert r.status_code == 200 and r.json()["precio"] == 25000

    # delete = soft (activo False), el cliente deja de verlo
    assert client.delete(f"/productos/{prod['id']}", headers=admin_headers).status_code == 204
    ids_cliente = [p["id"] for p in client.get("/productos/", headers=cliente.headers).json()]
    assert prod["id"] not in ids_cliente
    # staff lo sigue viendo con solo_activos=false
    ids_staff = [p["id"] for p in client.get("/productos/?solo_activos=false", headers=admin_headers).json()]
    assert prod["id"] in ids_staff


def test_cliente_siempre_ve_solo_activos(client, admin_headers, cliente):
    prod = _crear_producto(client, admin_headers)
    client.delete(f"/productos/{prod['id']}", headers=admin_headers)
    # aunque pida solo_activos=false, al cliente se le fuerza activos
    ids = [p["id"] for p in client.get("/productos/?solo_activos=false", headers=cliente.headers).json()]
    assert prod["id"] not in ids


def test_crear_producto_cliente_403(client, cliente):
    r = client.post("/productos/", json={"nombre": "X", "precio": 1, "stock": 1}, headers=cliente.headers)
    assert r.status_code == 403


# ── Ventas ─────────────────────────────────────────────────────


def test_venta_descuenta_stock_y_no_espeja_movimiento(client, admin_headers, db_session):
    prod = _crear_producto(client, admin_headers, precio=20000, stock=10)
    antes_mov = db_session.query(models.MovimientoFinanciero).count()

    r = _vender(client, admin_headers, prod["id"], cantidad=3)
    assert r.status_code == 201
    body = r.json()
    assert body["total"] == 60000
    assert body["precio_unitario"] == 20000

    p = db_session.query(models.Producto).filter_by(id=prod["id"]).first()
    assert p.stock == 7
    # regresión commit 8b8ff01: la venta NO crea MovimientoFinanciero
    assert db_session.query(models.MovimientoFinanciero).count() == antes_mov


def test_venta_stock_insuficiente_400(client, admin_headers, db_session):
    prod = _crear_producto(client, admin_headers, stock=2)
    assert _vender(client, admin_headers, prod["id"], cantidad=3).status_code == 400
    p = db_session.query(models.Producto).filter_by(id=prod["id"]).first()
    assert p.stock == 2  # intacto


def test_venta_producto_inactivo_400(client, admin_headers):
    prod = _crear_producto(client, admin_headers)
    client.delete(f"/productos/{prod['id']}", headers=admin_headers)
    assert _vender(client, admin_headers, prod["id"]).status_code == 400


def test_venta_producto_inexistente_404(client, admin_headers):
    assert _vender(client, admin_headers, 999999).status_code == 404


def test_venta_cantidad_invalida_422(client, admin_headers):
    prod = _crear_producto(client, admin_headers)
    assert _vender(client, admin_headers, prod["id"], cantidad=0).status_code == 422


def test_venta_cliente_403(client, admin_headers, cliente):
    prod = _crear_producto(client, admin_headers)
    assert _vender(client, cliente.headers, prod["id"]).status_code == 403


def test_listar_ventas(client, admin_headers):
    prod = _crear_producto(client, admin_headers)
    _vender(client, admin_headers, prod["id"])
    r = client.get("/ventas/", headers=admin_headers)
    assert r.status_code == 200 and len(r.json()) == 1


# ── Balance ────────────────────────────────────────────────────


def test_balance_sin_doble_conteo(client, admin_headers, cliente, db_session):
    # 1 pago de membresía (100k) + 1 venta (20k) + 1 ingreso manual (10k) + 1 egreso (5k)
    plan = db_session.query(models.Plan).filter_by(nombre="1 Mes").first()
    client.post(
        "/pagos/",
        json={"usuario_id": cliente.user.id, "plan_id": plan.id, "monto": 100000, "metodo_pago": "efectivo"},
        headers=admin_headers,
    )
    prod = _crear_producto(client, admin_headers, precio=20000, stock=5)
    _vender(client, admin_headers, prod["id"])
    assert _mov(client, admin_headers, "ingreso", 10000).status_code == 201
    assert _mov(client, admin_headers, "egreso", 5000, categoria="servicios").status_code == 201

    r = client.get("/finanzas/balance", headers=admin_headers)
    assert r.status_code == 200
    b = r.json()
    assert b["ingresos_total"] == 130000  # exactamente la suma, sin espejos
    assert b["total_membresias"] == 100000
    assert b["total_tienda"] == 20000
    assert b["egresos_total"] == 5000
    assert b["balance_neto"] == 125000
    assert b["ingresos_por_categoria"]["ingreso_varios"] == 10000
    assert b["egresos_por_categoria"]["servicios"] == 5000


def test_balance_filtro_fechas_excluye(client, admin_headers):
    _mov(client, admin_headers, "ingreso", 10000)
    # El movimiento se crea con datetime.utcnow(); calcular "mañana" desde esa
    # misma base para que el test no dependa de la hora local (después de las
    # 7 PM Bogotá, utcnow ya es el día siguiente).
    manana = (datetime.utcnow() + timedelta(days=1)).date().isoformat()
    r = client.get(f"/finanzas/balance?fecha_desde={manana}", headers=admin_headers)
    assert r.json()["ingresos_total"] == 0


def test_balance_solo_admin(client, coach, cliente):
    assert client.get("/finanzas/balance", headers=coach.headers).status_code == 403
    assert client.get("/finanzas/balance", headers=cliente.headers).status_code == 403


# ── Movimientos ────────────────────────────────────────────────


def test_movimientos_unifica_tres_fuentes(client, admin_headers, cliente, db_session):
    plan = db_session.query(models.Plan).filter_by(nombre="1 Semana").first()
    client.post(
        "/pagos/",
        json={"usuario_id": cliente.user.id, "plan_id": plan.id, "monto": 35000, "metodo_pago": "efectivo"},
        headers=admin_headers,
    )
    prod = _crear_producto(client, admin_headers)
    _vender(client, admin_headers, prod["id"])
    _mov(client, admin_headers, "egreso", 1000, categoria="servicios")

    r = client.get("/finanzas/movimientos", headers=admin_headers)
    assert r.status_code == 200
    items = r.json()["items"]
    prefijos = {i["id"].split("_")[0] for i in items}
    assert prefijos == {"pago", "venta", "mov"}
    # solo los manuales son eliminables
    for i in items:
        assert i["es_eliminable"] == i["id"].startswith("mov_")


def test_movimientos_filtro_egreso(client, admin_headers, cliente, db_session):
    plan = db_session.query(models.Plan).first()
    client.post(
        "/pagos/",
        json={"usuario_id": cliente.user.id, "plan_id": plan.id, "monto": 35000, "metodo_pago": "efectivo"},
        headers=admin_headers,
    )
    _mov(client, admin_headers, "egreso", 1000, categoria="servicios")
    body = client.get("/finanzas/movimientos?tipo=egreso", headers=admin_headers).json()
    assert body["total"] == 1
    assert len(body["items"]) == 1 and body["items"][0]["tipo"] == "egreso"


# ── Buscador y filtros del historial ───────────────────────────


def test_buscador_encuentra_por_concepto_y_por_cliente(client, admin_headers, cliente, db_session):
    plan = db_session.query(models.Plan).first()
    _pagar(client, admin_headers, cliente.user.id, plan.id)
    _mov(client, admin_headers, "egreso", 1000, categoria="renta", concepto="Alquiler del local")

    def buscar(q):
        return client.get(f"/finanzas/movimientos?q={q}", headers=admin_headers).json()

    por_concepto = buscar("alquiler")
    assert por_concepto["total"] == 1
    assert "Alquiler" in por_concepto["items"][0]["concepto"]

    # El nombre del cliente no es una columna del movimiento: se arma en el helper, y
    # por eso el buscador tiene que pegar contra el texto ya construido.
    por_cliente = buscar(cliente.user.nombre[:8])
    assert por_cliente["total"] >= 1
    assert all(cliente.user.nombre[:8] in m["concepto"] for m in por_cliente["items"])


def test_buscador_ignora_los_acentos(client, admin_headers):
    """Quien busca no tiene por qué saber con qué acento se cargó el concepto."""
    _mov(client, admin_headers, "egreso", 5000, categoria="nomina", concepto="Nómina de agosto")

    body = client.get("/finanzas/movimientos?q=nomina", headers=admin_headers).json()
    assert body["total"] == 1 and body["items"][0]["concepto"] == "Nómina de agosto"


def test_filtro_por_categoria(client, admin_headers, cliente, db_session):
    plan = db_session.query(models.Plan).first()
    _pagar(client, admin_headers, cliente.user.id, plan.id)
    _mov(client, admin_headers, "egreso", 1000, categoria="renta")
    _mov(client, admin_headers, "egreso", 2000, categoria="servicios")

    body = client.get("/finanzas/movimientos?categoria=renta", headers=admin_headers).json()
    assert body["total"] == 1 and body["items"][0]["categoria"] == "renta"

    # "mensualidad" tiene que traer los pagos, que no son movimientos manuales.
    solo_membresias = client.get("/finanzas/movimientos?categoria=mensualidad", headers=admin_headers).json()
    assert solo_membresias["total"] == 1
    assert solo_membresias["items"][0]["fuente"] == "pago_membresia"


def test_filtro_por_plan_deja_fuera_ventas_y_manuales(client, admin_headers, cliente, db_session):
    """Un plan solo puede tener pagos de membresía detrás: una venta de tienda o un
    egreso de renta no pertenecen a ningún plan y colarlos haría mentir al total."""
    plan_a = db_session.query(models.Plan).filter_by(nombre="1 Semana").first()
    plan_b = db_session.query(models.Plan).filter(models.Plan.id != plan_a.id).first()
    _pagar(client, admin_headers, cliente.user.id, plan_a.id, monto=10000)
    _pagar(client, admin_headers, cliente.user.id, plan_b.id, monto=90000)
    _vender(client, admin_headers, _crear_producto(client, admin_headers)["id"])
    _mov(client, admin_headers, "egreso", 1000, categoria="renta")

    body = client.get(f"/finanzas/movimientos?plan_id={plan_a.id}", headers=admin_headers).json()
    assert body["total"] == 1
    assert body["items"][0]["monto"] == 10000
    assert body["items"][0]["fuente"] == "pago_membresia"


def test_paginacion_no_cambia_el_total(client, admin_headers):
    for i in range(5):
        _mov(client, admin_headers, "egreso", 100 + i, categoria="otros")

    p1 = client.get("/finanzas/movimientos?skip=0&limit=2", headers=admin_headers).json()
    p2 = client.get("/finanzas/movimientos?skip=2&limit=2", headers=admin_headers).json()

    assert p1["total"] == p2["total"] == 5      # el total es del período, no de la página
    assert len(p1["items"]) == len(p2["items"]) == 2
    assert {m["id"] for m in p1["items"]}.isdisjoint({m["id"] for m in p2["items"]})


# ── Estadísticas: qué se vendió ────────────────────────────────


def test_estadisticas_agrupa_por_plan(client, admin_headers, cliente, crear_usuario, db_session):
    plan_a = db_session.query(models.Plan).filter_by(nombre="1 Semana").first()
    plan_b = db_session.query(models.Plan).filter(models.Plan.id != plan_a.id).first()
    otro = crear_usuario("cliente")

    _pagar(client, admin_headers, cliente.user.id, plan_a.id, monto=10000)
    _pagar(client, admin_headers, otro.user.id, plan_a.id, monto=10000)
    _pagar(client, admin_headers, cliente.user.id, plan_b.id, monto=5000)

    datos = client.get("/finanzas/estadisticas", headers=admin_headers).json()
    por_nombre = {p["nombre"]: p for p in datos["planes"]}

    assert por_nombre[plan_a.nombre]["cantidad"] == 2
    assert por_nombre[plan_a.nombre]["total"] == 20000
    assert por_nombre[plan_b.nombre]["cantidad"] == 1
    assert datos["total_planes"] == 25000
    # Ordenado por lo que más facturó, que es como se lee la tabla.
    assert datos["planes"][0]["nombre"] == plan_a.nombre
    assert por_nombre[plan_a.nombre]["porcentaje"] == 80.0


def test_estadisticas_cuenta_los_personalizados_aparte(client, admin_headers, cliente):
    """`plan_id` es NULL en un pago personalizado: con un join normal desaparecería del
    desglose, y es ingreso real que ya está contado en la tarjeta de Membresías."""
    client.post(
        "/pagos/directo/",
        json={"usuario_id": cliente.user.id, "duracion_dias": 20, "monto": 40000, "metodo_pago": "efectivo"},
        headers=admin_headers,
    )
    datos = client.get("/finanzas/estadisticas", headers=admin_headers).json()
    personalizado = next(p for p in datos["planes"] if p["nombre"] == "Personalizado")
    assert personalizado["plan_id"] is None
    assert personalizado["cantidad"] == 1 and personalizado["total"] == 40000


def test_estadisticas_agrupa_por_producto(client, admin_headers):
    prod = _crear_producto(client, admin_headers, precio=20000, stock=10, nombre="Proteína")
    _vender(client, admin_headers, prod["id"], cantidad=2)
    _vender(client, admin_headers, prod["id"], cantidad=1)

    datos = client.get("/finanzas/estadisticas", headers=admin_headers).json()
    fila = next(p for p in datos["productos"] if p["nombre"] == "Proteína")
    assert fila["unidades"] == 3
    assert fila["total"] == 60000


def test_estadisticas_solo_admin(client, coach, cliente):
    assert client.get("/finanzas/estadisticas", headers=coach.headers).status_code == 403
    assert client.get("/finanzas/estadisticas", headers=cliente.headers).status_code == 403


# ── Zona horaria: la ventana es el día de Bogotá, no el de UTC ──


def test_ventana_es_el_dia_de_bogota_no_el_de_utc(client, admin_headers, cliente, db_session):
    """Las columnas se guardan naive en UTC y el filtro recibe días de Bogotá.

    Un caso a mediodía no prueba nada: el bug vive entre las 19:00 y la medianoche, que
    en UTC ya son del día siguiente. Se atacan las dos puntas — lo de hoy a las 21:00
    tiene que entrar, y lo de anoche a las 20:00 tiene que quedar afuera.
    """
    from datetime import time

    from fechas import a_utc, hoy_bogota

    hoy = hoy_bogota()
    ayer = hoy - timedelta(days=1)
    plan = db_session.query(models.Plan).first()

    db_session.add_all([
        models.Pago(usuario_id=cliente.user.id, plan_id=plan.id, monto=111, metodo_pago="efectivo",
                    fecha_pago=a_utc(datetime.combine(hoy, time(21, 0)))),
        models.Pago(usuario_id=cliente.user.id, plan_id=plan.id, monto=222, metodo_pago="efectivo",
                    fecha_pago=a_utc(datetime.combine(ayer, time(20, 0)))),
    ])
    db_session.commit()

    body = client.get(
        f"/finanzas/movimientos?fecha_desde={hoy.isoformat()}&fecha_hasta={hoy.isoformat()}",
        headers=admin_headers,
    ).json()
    montos = {m["monto"] for m in body["items"]}

    assert 111 in montos, "la noche de hoy se perdía: 21:00 Bogotá ya es mañana en UTC"
    assert 222 not in montos, "la noche de ayer se colaba en el día de hoy"


def test_eliminar_movimiento_manual(client, admin_headers, db_session):
    _mov(client, admin_headers, "ingreso", 500)
    mov = db_session.query(models.MovimientoFinanciero).first()
    assert client.delete(f"/finanzas/movimientos/{mov.id}", headers=admin_headers).status_code == 204


def test_eliminar_movimiento_no_manual_404(client, admin_headers, db_session):
    db_session.add(models.MovimientoFinanciero(
        tipo=models.TipoMovimiento.INGRESO, concepto="legacy", categoria="mensualidad",
        monto=100, fecha=datetime.utcnow(), fuente="pago_directo",
    ))
    db_session.commit()
    mov = db_session.query(models.MovimientoFinanciero).first()
    assert client.delete(f"/finanzas/movimientos/{mov.id}", headers=admin_headers).status_code == 404


def test_movimiento_monto_invalido_422(client, admin_headers):
    assert _mov(client, admin_headers, "ingreso", 0).status_code == 422
    assert _mov(client, admin_headers, "regalo", 100).status_code == 422


# ── Exportar Excel ─────────────────────────────────────────────


def test_exportar_excel_admin(client, admin_headers):
    """Tres hojas: Resumen, Ingresos y Egresos, cada una con su total al pie."""
    import io

    from openpyxl import load_workbook

    _mov(client, admin_headers, "ingreso", 50000)
    _mov(client, admin_headers, "egreso", 20000, categoria="nomina")

    r = client.get("/finanzas/exportar-excel", headers=admin_headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "finanzas_jainsportbox_" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Resumen", "Ingresos", "Egresos"]

    resumen = {f[0]: f[1] for f in wb["Resumen"].iter_rows(values_only=True) if f[0]}
    assert resumen["Total ingresos"] == 50000
    assert resumen["Egresos"] == 20000
    assert resumen["Balance neto"] == 30000

    assert resumen["Nómina"] == 20000   # desglose de egresos por categoría

    # Cada hoja lleva solo su tipo, con montos positivos y una fila TOTAL al pie.
    ing = list(wb["Ingresos"].iter_rows(values_only=True))
    assert ing[0][0] == "Fecha" and ing[0][-1] == "Monto" and "Origen" in ing[0]
    assert [f[-1] for f in ing[1:] if f[-1] is not None] == [50000, 50000]  # fila + TOTAL

    egr = list(wb["Egresos"].iter_rows(values_only=True))
    assert "Origen" not in egr[0]   # en egresos sería siempre "Registro manual"
    assert [f[-1] for f in egr[1:] if f[-1] is not None] == [20000, 20000]


def test_exportar_excel_sin_egresos_muestra_la_seccion(client, admin_headers):
    """Sin egresos la sección igual aparece: si se omitiera, no se sabría si no
    hubo gastos o si el export falló."""
    import io

    from openpyxl import load_workbook

    _mov(client, admin_headers, "ingreso", 10000)

    r = client.get("/finanzas/exportar-excel", headers=admin_headers)
    resumen = {
        f[0]: f[1]
        for f in load_workbook(io.BytesIO(r.content))["Resumen"].iter_rows(values_only=True)
        if f[0]
    }
    assert "Egresos por categoría" in resumen
    assert resumen["Sin egresos en el período"] == 0


def test_exportar_excel_no_admin_403(client, coach, cliente):
    assert client.get("/finanzas/exportar-excel", headers=coach.headers).status_code == 403
    assert client.get("/finanzas/exportar-excel", headers=cliente.headers).status_code == 403


